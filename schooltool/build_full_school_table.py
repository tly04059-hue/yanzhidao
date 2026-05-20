import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
MAIN_XLSX = BASE / "研招网MBA、MPA、云贵川渝清单整理.xlsx"
FULL_CSV = BASE / "mba_full_database_final_schoolid_full.csv"
SCHOOLS_CSV = BASE / "mba_schools_schoolid_full.csv"
SCORE_CSV = BASE / "mba_score_band_stats_schoolid_full.csv"
OUTPUT = BASE / "outputs" / "全量院校表_研招网主表补全.xlsx"


def read_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def norm_college(value) -> str:
    value = clean_text(value)
    value = re.sub(r"^\s*[（(][^）)]*[）)]\s*", "", value)
    return value.replace(" ", "")


def norm_direction(value) -> str:
    value = clean_text(value)
    value = re.sub(r"[（(][^）)]*[）)]", "", value)
    value = re.sub(r"\s+", "", value)
    return value


def first_non_empty(values) -> str:
    for value in values:
        value = clean_text(value)
        if value and value.lower() != "nan":
            return value
    return ""


def join_unique(values, limit=12) -> str:
    seen = []
    for value in values:
        value = clean_text(value)
        if value and value not in seen:
            seen.append(value)
    if len(seen) > limit:
        return "；".join(seen[:limit]) + f"；等{len(seen)}项"
    return "；".join(seen)


def score_summary(score_df: pd.DataFrame) -> pd.DataFrame:
    if score_df.empty:
        return pd.DataFrame()

    def summarize_group(group: pd.DataFrame) -> pd.Series:
        group = group.sort_values(["年份", "项目类型", "分数段"], na_position="last")
        parts = []
        for year, year_group in group.groupby("年份", dropna=False):
            rows = []
            for _, row in year_group.iterrows():
                band = clean_text(row.get("分数段"))
                admitted = clean_text(row.get("拟录人数"))
                retest = clean_text(row.get("进入复试人数"))
                rate = clean_text(row.get("分段录取率"))
                label = clean_text(row.get("项目类型"))
                rows.append(f"{label}{band}:拟录{admitted}/复试{retest}/录取率{rate}")
            parts.append(f"{year}：" + "，".join(rows[:20]))
        return pd.Series(
            {
                "分数段年份": join_unique(group["年份"]),
                "录取分析标题": join_unique(group["录取分析标题"]),
                "总成绩计算方式": join_unique(group["总成绩计算方式"], limit=5),
                "分数段统计": "；".join(parts[:8]),
                "分数段来源网站": join_unique(group["来源网站"]),
                "分数段更新时间": join_unique(group["更新时间"], limit=5),
            }
        )

    return score_df.groupby("项目ID", dropna=False).apply(summarize_group).reset_index()


def best_full_match(row: pd.Series, full_candidates: pd.DataFrame) -> pd.Series | None:
    school = row["_学校名称"]
    project = row["_项目"]
    college = row["_院系_norm"]
    study = row["_学习方式"]
    direction = row["_方向_norm"]

    candidates = full_candidates[
        (full_candidates["_学校"] == school) & (full_candidates["_请求专业类型"] == project)
    ]
    if candidates.empty:
        return None

    levels = [
        candidates[
            (candidates["_学院_norm"] == college)
            & (candidates["_项目类型"] == study)
            & (
                (candidates["_项目名称_norm"] == direction)
                | (candidates["_标准项目名称_norm"] == direction)
                | (candidates["_方向_norm"] == direction)
            )
        ],
        candidates[(candidates["_学院_norm"] == college) & (candidates["_项目类型"] == study)],
        candidates[candidates["_学院_norm"] == college],
        candidates[candidates["_项目类型"] == study],
        candidates,
    ]
    for level in levels:
        if not level.empty:
            return level.iloc[0]
    return None


def main() -> None:
    main_df = pd.read_excel(MAIN_XLSX, sheet_name="全量清单", dtype=str).fillna("")
    full_df = read_csv(FULL_CSV)
    schools_df = read_csv(SCHOOLS_CSV)
    score_df = read_csv(SCORE_CSV)

    main_columns = list(main_df.columns)

    for df in [main_df, full_df, schools_df, score_df]:
        df.columns = [clean_text(c) for c in df.columns]

    main_df["_来源行号"] = range(2, len(main_df) + 2)
    main_df["_学校名称"] = main_df["学校名称"].map(clean_text)
    main_df["_项目"] = main_df["项目"].map(clean_text)
    main_df["_院系_norm"] = main_df["院系"].map(norm_college)
    main_df["_方向_norm"] = main_df["方向"].map(norm_direction)
    main_df["_学习方式"] = main_df["学习方式"].map(clean_text)

    full_df["_学校"] = full_df["学校"].map(clean_text)
    full_df["_请求专业类型"] = full_df["请求专业类型"].map(clean_text)
    full_df["_学院_norm"] = full_df["学院"].map(norm_college)
    full_df["_项目名称_norm"] = full_df["项目名称"].map(norm_direction)
    full_df["_标准项目名称_norm"] = full_df["标准项目名称"].map(norm_direction)
    full_df["_方向_norm"] = full_df["方向"].map(norm_direction)
    full_df["_项目类型"] = full_df["项目类型"].map(clean_text)

    score_agg = score_summary(score_df)
    full_enriched = full_df.merge(score_agg, on="项目ID", how="left")
    schools_pick = schools_df.drop_duplicates("项目ID", keep="first").add_prefix("学校概览_")
    full_enriched = full_enriched.merge(
        schools_pick, left_on="项目ID", right_on="学校概览_项目ID", how="left"
    )

    supplement_cols = [
        "项目ID",
        "请求专业类型",
        "学校",
        "学院",
        "学校简介",
        "院校分数线类别",
        "项目名称",
        "标准项目名称",
        "项目类型",
        "学制",
        "上课地点",
        "上课方式",
        "城市",
        "院校性质",
        "官网",
        "是否提前面试",
        "是否接受调剂",
        "是否接受大专报考",
        "2025分数线",
        "2024分数线",
        "2023分数线",
        "2022分数线",
        "最新分数线年份",
        "最新分数线",
        "总分",
        "英语",
        "综合",
        "复试信息",
        "录取分析",
        "更新时间",
        "学校概览_大学ID",
        "学校概览_详情页",
        "学校概览_大学名称",
        "学校概览_学院专业列表",
        "学校概览_学院招生年份",
        "分数段年份",
        "录取分析标题",
        "总成绩计算方式",
        "分数段统计",
        "分数段来源网站",
        "分数段更新时间",
    ]

    rows = []
    matched_project_ids = set()
    for _, row in main_df.iterrows():
        match = best_full_match(row, full_enriched)
        out = {col: row.get(col, "") for col in main_columns}
        out["数据来源"] = "研招网主表；MBA大师详情库补全" if match is not None else "研招网主表"
        out["匹配状态"] = "已匹配详情库" if match is not None else "未匹配详情库"
        out["主表来源行号"] = row["_来源行号"]
        for col in supplement_cols:
            out[col] = clean_text(match.get(col, "")) if match is not None else ""
        if match is not None:
            matched_project_ids.add(clean_text(match.get("项目ID")))
        rows.append(out)

    covered_schools = set(main_df["_学校名称"])
    extra_full = full_enriched[~full_enriched["_学校"].isin(covered_schools)].copy()
    for _, row in extra_full.iterrows():
        out = {col: "" for col in main_columns}
        out.update(
            {
                "地区": "",
                "项目": clean_text(row.get("请求专业类型")),
                "学校代码": "",
                "学校名称": clean_text(row.get("学校")),
                "学校所在地": clean_text(row.get("城市")),
                "学校标签": first_non_empty([row.get("院校性质"), row.get("学校概览_性质")]),
                "院系": clean_text(row.get("学院")),
                "考试类型": "",
                "专业代码": "",
                "学习方式": clean_text(row.get("项目类型")),
                "学费": clean_text(row.get("学费")),
                "方向": first_non_empty([row.get("项目名称"), row.get("标准项目名称"), row.get("方向")]),
                "招生人数": clean_text(row.get("招生人数")),
                "考试科目": "",
                "去年分数线": clean_text(row.get("2024分数线")),
                "今年分数线": clean_text(row.get("2025分数线")),
                "录取情况": clean_text(row.get("录取分析")),
                "录取率": "",
                "退役计划": "",
                "少民计划": "",
                "备注": clean_text(row.get("复试信息")),
                "学校Logo": "",
            }
        )
        out["数据来源"] = "MBA大师详情库补充"
        out["匹配状态"] = "详情库新增院校/项目"
        out["主表来源行号"] = ""
        for col in supplement_cols:
            out[col] = clean_text(row.get(col, ""))
        rows.append(out)

    merged = pd.DataFrame(rows)
    final_cols = main_columns + ["数据来源", "匹配状态", "主表来源行号"] + supplement_cols
    merged = merged[final_cols].drop_duplicates()

    analysis = pd.DataFrame(
        [
            ["研招网MBA、MPA、云贵川渝清单整理.xlsx", "全量清单", len(main_df), len(main_columns), main_df["学校名称"].nunique(), "云贵川渝 MBA/MPA 目录；字段最贴近研招网招生简章，作为主表字段"],
            ["mba_full_database_final_schoolid_full.csv", "详情库", len(full_df), len(full_df.columns) - 8, full_df["学校"].nunique(), "全国 MBA/MPA 项目详情；含项目ID、学费、学制、分数线、复试和录取分析"],
            ["mba_schools_schoolid_full.csv", "学校概览", len(schools_df), len(schools_df.columns), schools_df["学校"].nunique(), "学校-学院-专业概览；含大学ID、详情页、性质、学院专业列表"],
            ["mba_score_band_stats_schoolid_full.csv", "分数段统计", len(score_df), len(score_df.columns), score_df["学校"].nunique(), "按项目ID记录年份、分数段、复试人数、拟录人数和分段录取率；已聚合后并入全量表"],
        ],
        columns=["文件", "数据层级", "行数", "字段数", "学校数", "特点"],
    )

    merge_note = pd.DataFrame(
        [
            ["主表保留", "Excel 全量清单 181 行全部保留，原始 22 个字段位于最左侧。"],
            ["补充范围", "CSV 详情库中不属于主表 45 所学校的记录作为新增院校/项目追加。"],
            ["匹配规则", "主表记录按 学校+项目+院系+学习方式+方向 优先匹配详情库；不成功时逐级放宽到 学校+项目。"],
            ["分数段处理", "score_band_stats 按 项目ID 聚合成 分数段年份、总成绩计算方式、分数段统计 等字段，避免一对多导致行数膨胀。"],
            ["输出规模", f"合并后 {len(merged)} 行、{len(merged.columns)} 列。"],
        ],
        columns=["项目", "说明"],
    )

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="全量院校表", index=False)
        analysis.to_excel(writer, sheet_name="文件分析", index=False)
        merge_note.to_excel(writer, sheet_name="合并说明", index=False)

    wb = load_workbook(OUTPUT)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, col_cells in enumerate(ws.columns, start=1):
            header = clean_text(ws.cell(row=1, column=idx).value)
            sample_lengths = [len(clean_text(c.value)) for c in list(col_cells)[:200]]
            width = min(max(max(sample_lengths + [len(header)]) + 2, 10), 42)
            ws.column_dimensions[get_column_letter(idx)].width = width
        if ws.title == "全量院校表":
            for idx in range(1, len(main_columns) + 1):
                ws.cell(row=1, column=idx).fill = sub_fill
                ws.cell(row=1, column=idx).font = Font(color="000000", bold=True)
    wb.save(OUTPUT)

    print(f"saved={OUTPUT}")
    print(f"rows={len(merged)} cols={len(merged.columns)}")
    print(f"main_rows={len(main_df)} extra_full_rows={len(extra_full)} matched_main_rows={(merged['匹配状态'] == '已匹配详情库').sum()}")


if __name__ == "__main__":
    main()
