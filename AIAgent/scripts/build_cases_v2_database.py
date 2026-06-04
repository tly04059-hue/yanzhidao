from __future__ import annotations

import html
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_party_duplicate_rewrite_draft import feedback as draft_feedback


ROOT = Path(__file__).resolve().parents[2]
PARTY_SOURCE = ROOT / "miniapp" / "dx-cases.json"
EXAM_SOURCE = ROOT / "studentCases" / "sourceData" / "管综已上岸学员案例信息汇总表.xlsx"
DB_DIR = ROOT / "studentCases" / "cases-v2-db"
REPORT = ROOT / "studentCases" / "reports" / "cases-v2-review.html"
MINIAPP_PUBLIC_DATA = ROOT / "miniapp" / "src" / "data" / "cases-v2-public.json"
ANALYTICS_DASHBOARD = ROOT / "studentCases" / "reports" / "cases-v2-analytics-dashboard.html"

NARRATIVE_KEYS = [
    "key_quote",
    "narrative_choose",
    "reflection",
    "story_summary",
    "study_method",
    "turning_point",
]
NARRATIVE_WEIGHTS = {
    "key_quote": 3.0,
    "narrative_choose": 2.0,
    "reflection": 2.5,
    "story_summary": 2.0,
    "study_method": 1.5,
    "turning_point": 1.5,
}
PUBLIC_PARTY_FIELDS = [
    ("key_quote", "报考动机"),
    ("reflection", "真实反馈"),
    ("study_method", "备考方法"),
]
HIDDEN_TAGS = {"其他", "M8-resync", "system_unknown_storied", "心得型"}
PARTY_CASE_IDS_TO_DROP = {"C123"}
PARTY_TARGET_OVERRIDES = {
    "C022": "四川党校 · 政治学",
    "C025": "四川党校 · 经济学",
    "C112": "四川党校 · 政治学",
    "C122": "四川党校 · 政治学",
}
DISTRICT_UNIT_PATTERN = re.compile(
    r"(青羊区|锦江区|武侯区|金牛区|成华区|高新区|天府新区|双流区|温江区|郫都区|新都区|龙泉驿区|"
    r"[^，,；;、\s]{1,6}区|[^，,；;、\s]{1,6}县|[^，,；;、\s]{1,6}市)委党校"
)
IMPRECISE_TARGET_PATTERN = re.compile(
    r"(及附近|名校|好上岸|第一优先|川渝院校|重庆院校|医科类院校|医学类院校|类院校|成都985|重庆985|新疆|学费低|离雅安|不详)"
)
AGE_RANGE_PATTERN = re.compile(r"(\d{2})\s*-\s*(\d{2})\s*岁")


def clean(value: Any) -> str:
    return str(value or "").strip()


def normalize_spaces(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value))


def display_age(value: Any) -> str:
    text = clean(value).replace("岁段", "岁")
    match = re.fullmatch(r"(\d{2})\s*-\s*(\d{2})\s*岁?", text)
    if match:
        start, end = (int(match.group(1)), int(match.group(2)))
        return f"约 {(start + end) // 2} 岁"
    return normalize_age_ranges_in_text(text)


def normalize_age_ranges_in_text(value: Any) -> str:
    text = clean(value).replace("岁段", "岁")
    text = AGE_RANGE_PATTERN.sub(lambda match: f"约 {(int(match.group(1)) + int(match.group(2))) // 2} 岁", text)
    return re.sub(r"约\s*(\d{1,2})\s*岁", r"约 \1 岁", text)


def escape(value: Any) -> str:
    return html.escape(clean(value), quote=True)


def clean_display_text(value: Any) -> str:
    text = clean(value)
    text = text.replace("M8-resync", "")
    text = text.replace("其他", "")
    text = normalize_age_ranges_in_text(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def visible_source_value(value: Any) -> Any:
    if isinstance(value, list):
        result = []
        for item in value:
            cleaned = clean_display_text(item)
            if cleaned and cleaned not in HIDDEN_TAGS:
                result.append(cleaned)
        return result
    if not isinstance(value, str):
        return value
    cleaned = clean_display_text(value)
    if not cleaned or cleaned in HIDDEN_TAGS:
        return ""
    return cleaned


def has_visible_source_value(value: Any) -> bool:
    if isinstance(value, list):
        return bool(value)
    return bool(clean(value))


def party_source_payload(case: dict[str, Any], target: str) -> dict[str, Any]:
    payload = dict(case)
    payload["chosen_school"] = target
    if "age_band" in payload:
        payload["age_band"] = display_age(payload["age_band"])
    if "age_concrete" in payload:
        payload["age_concrete"] = display_age(payload["age_concrete"])
    for key, value in list(payload.items()):
        payload[key] = visible_source_value(value)
    return payload


def clean_party_visible_text(value: Any) -> str:
    text = clean_display_text(value)
    text = re.sub(r"(四川省委党校|中共四川省委党校|重庆市委党校|中共重庆市委党校|四川党校|重庆党校|省委党校|市委党校)", "", text)
    text = text.replace("党校在职研究生", "在职研究生")
    text = text.replace("党校研究生", "在职研究生")
    text = text.replace("考党校", "考在职研")
    text = text.replace("报考党校", "报考在职研")
    text = text.replace("备考党校", "备考在职研")
    text = text.replace("党校", "")
    text = re.sub(r"\s*·\s*·\s*", " · ", text)
    text = re.sub(r"(^|\s)[·，,、；;：:]\s*", r"\1", text)
    text = re.sub(r"\s*[·，,、；;：:]$", "", text)
    return re.sub(r"\s+", " ", text).strip()


def sanitize_exam_private_text(value: Any) -> str:
    text = clean(value)
    text = DISTRICT_UNIT_PATTERN.sub("相关系统", text)
    text = re.sub(r"(青羊区|锦江区|武侯区|金牛区|成华区|高新区|天府新区|双流区|温江区|郫都区|新都区|龙泉驿区)", "当地", text)
    text = text.replace("党校系统", "相关系统")
    text = text.replace("党校单证", "单证")
    text = text.replace("党校", "在职研")
    return text.strip(" ，,；;、")


def compact_tags(*groups: list[str] | None, hide_party: bool = False) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for group in groups:
        for item in group or []:
            tag = clean_display_text(item)
            if not tag or tag in HIDDEN_TAGS:
                continue
            if hide_party:
                tag = clean_party_visible_text(tag)
            if tag and tag not in seen:
                seen.add(tag)
                result.append(tag)
    return result


def deduped_narratives(case: dict[str, Any]) -> list[tuple[str, str]]:
    kept: list[tuple[str, str]] = []
    for key in NARRATIVE_KEYS:
        text = clean(case.get(key))
        if not text:
            continue
        replaced = False
        for idx, (_, existing_text) in enumerate(kept):
            if text == existing_text or existing_text in text or text in existing_text:
                if len(text) > len(existing_text):
                    kept[idx] = (key, text)
                replaced = True
                break
        if not replaced:
            kept.append((key, text))
    return kept


def party_richness(case: dict[str, Any]) -> int:
    score = 0.0
    for key, text in deduped_narratives(case):
        score += min(len(text), 180) * NARRATIVE_WEIGHTS[key]
    return round(score)


def party_ai_info(case: dict[str, Any]) -> dict[str, Any]:
    core_values = [clean(case.get(key)) for key in ["key_quote", "narrative_choose", "reflection"]]
    is_duplicate = all(core_values) and len(set(core_values)) == 1
    if not is_duplicate:
        return {
            "ai_fill_level": "none",
            "ai_filled_fields": [],
            "ai_fill_note": "原始字段可直接展示，未使用 AI 补充。",
            "ai_sort_weight": 0,
        }
    return {
        "ai_fill_level": "partial",
        "ai_filled_fields": ["reflection"],
        "ai_fill_note": "真实反馈由 AI 根据该案例已有字段补充；选择原因不进入用户端展示。",
        "ai_sort_weight": 1,
    }


def party_display_narratives(case: dict[str, Any]) -> dict[str, str]:
    narratives = {key: clean(case.get(key)) for key in NARRATIVE_KEYS}
    if party_ai_info(case)["ai_fill_level"] == "partial":
        narratives["reflection"] = draft_feedback(case)
    return narratives


def party_quality(score: int) -> str:
    if score >= 480:
        return "rich"
    if score >= 180:
        return "standard"
    return "compact"


def infer_program(target: str, fallback = "MPA") -> str:
    upper = clean(target).upper()
    for candidate in ["MEM", "MTA", "MBA", "MPA"]:
        if candidate in upper:
            return candidate
    return fallback


def normalize_target(source_target: str, fallback_program = "") -> tuple[str, str, str]:
    target = normalize_spaces(source_target)
    program = infer_program(target, fallback_program.upper() or "MPA")
    school = re.sub(program, "", target, flags=re.IGNORECASE).strip(" ·-—")
    school = normalize_spaces(school)
    normalized = f"{school} · {program}" if school else program
    precision = "precise" if school and not IMPRECISE_TARGET_PATTERN.search(school) else "broad"
    return normalized, program, precision


def exam_intent_and_admitted(row: dict[str, str]) -> tuple[str, str, str, str]:
    intent_raw = clean(row.get("意向院校") or row.get("原始上岸院校") or row.get("上岸院校"))
    admitted_raw = clean(row.get("上岸院校"))
    intent, program, precision = normalize_target(intent_raw, row.get("意向专业", ""))
    admitted, _, admitted_precision = normalize_target(admitted_raw, row.get("意向专业", ""))
    if admitted_precision == "precise":
        precision = admitted_precision
    return intent, admitted, program, precision


def age_band(age: str) -> str:
    raw = clean(age)
    try:
        value = int(float(raw))
    except ValueError:
        label = display_age(raw)
        if re.fullmatch(r"\d{2}\s*-\s*\d{2}", label):
            return f"{label.replace(' ', '')}岁"
        return label
    if value < 25:
        return "25岁以下"
    if value <= 30:
        return display_age("25-30岁")
    if value <= 35:
        return display_age("30-35岁")
    if value <= 40:
        return display_age("35-40岁")
    return "40岁以上"


def display_alias(name: str, gender: str) -> str:
    surname = clean(name)[:1] or "同"
    suffix = "女士" if clean(gender) == "女" else "先生" if clean(gender) == "男" else "同学"
    return f"{surname}{suffix}"


def infer_exam_system(job: str, motivation: str) -> str:
    text = f"{job} {motivation}"
    if re.search(r"人民银行|银行|金融", text):
        return "银行金融"
    if re.search(r"国企|央企", text):
        return "国央企"
    if re.search(r"教师|高职|学校|培训", text):
        return "教育系统"
    if re.search(r"公务员|事业编|体制内|机关|街道|交通|公路|精神文明", text):
        return "党政机关"
    return ""


def goal_tags(motivation: str) -> list[str]:
    text = clean(motivation)
    tags: list[str] = []
    if "遴选" in text:
        tags.append("遴选")
    if "晋升" in text or "向上发展" in text:
        tags.append("晋升")
    if "学历" in text:
        tags.append("学历提升")
    if "学习" in text or "充实自我" in text or "提升自己" in text:
        tags.append("自我提升")
    if "双证" in text:
        tags.append("双证")
    return tags or ["学历提升"]


def exam_risk(math_base: str, english_base: str, test_score: str, program: str) -> str:
    try:
        score = float(clean(test_score))
    except ValueError:
        score = 0
    if score and score < 50:
        return "测评分数偏低，基础补齐和择校保底需要优先处理。"
    if "未及格" in math_base or "薄弱" in math_base:
        return "数学基础是主要风险，择校不宜只看院校名气。"
    if "四级未过" in english_base or "未参加" in english_base:
        return "英语基础需要同步补齐，避免后期总分和单科线同时受影响。"
    if program == "MBA":
        return "MBA 更容易受学费、城市和上课方式约束。"
    if program == "MEM":
        return "MEM 需要兼顾项目适配度、复试要求和院校专业方向。"
    return "需要同时平衡目标院校、基础水平和备考时间。"


def exam_advice(program: str, precision: str, risk: str) -> str:
    if precision == "broad":
        return "先把意向院校、城市、学费和分数线拆开复盘，再沉淀成可复用择校口径。"
    if "数学" in risk or "测评" in risk:
        return "优先选择分数线稳定、招生人数相对充足、学费和上课方式匹配的院校，再安排基础补齐。"
    if program == "MBA":
        return "先锁定预算和上课地点，再比较复试压力与院校认可度。"
    if program == "MEM":
        return "优先确认专业方向和复试要求，再做冲刺与保底组合。"
    return "按专业方向和城市筛选，再结合分数线、学费、招生人数做保底和冲刺组合。"


def build_party_records() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    payload = json.loads(PARTY_SOURCE.read_text(encoding="utf-8"))
    source_cases = [item for item in payload.get("data", {}).get("cases", []) if clean(item.get("chosen_path")) == "A"]

    raw_records = []
    desensitized_records = []
    public_records = []
    for index, case in enumerate(source_cases, start=1):
        case_id = clean(case.get("case_id"))
        if case_id in PARTY_CASE_IDS_TO_DROP:
            continue
        target = PARTY_TARGET_OVERRIDES.get(case_id, clean(case.get("chosen_school")))
        ai_info = party_ai_info(case)
        narratives = party_display_narratives(case)
        richness = party_richness(case)
        sort_tuple = [ai_info["ai_sort_weight"], -richness, index]
        common = {
            "direction": "party",
            "source_dataset": str(PARTY_SOURCE.relative_to(ROOT)),
            "source_id": case_id,
            "display_order_source": index,
            "sort_tuple": sort_tuple,
            "richness_score": richness,
            "quality": party_quality(richness),
            "ai_fill_level": ai_info["ai_fill_level"],
            "ai_filled_fields": ai_info["ai_filled_fields"],
            "ai_fill_note": ai_info["ai_fill_note"],
            "human_review_status": "pending",
        }
        raw_records.append({
            **common,
            "privacy_version": "raw_internal",
            "privacy_note": "党校当前可用源本身已做基础脱敏；未脱敏原始 M8 不在本文件中。",
            "chosen_target": target,
            "source_payload": party_source_payload(case, target),
        })
        internal_fields = {
            "case_id": clean(case.get("case_id")),
            "display_name": clean(case.get("display_name")),
            "age": display_age(case.get("age_concrete") or case.get("age_band")),
            "region": clean(case.get("region")),
            "system": clean(case.get("system_chip") or case.get("system")),
            "position": clean(case.get("system") if clean(case.get("system")) != "system_unknown_storied" else case.get("unit_narrative")),
            "chosen_target": target,
            "outcome": clean(case.get("outcome")),
            "tags": compact_tags(case.get("tags") or [], case.get("reason_tags") or [], case.get("story_tags") or []),
            "fields": {
                "key_quote": clean_display_text(case.get("key_quote")),
                "narrative_choose": clean_display_text(case.get("narrative_choose")),
                "reflection": clean_display_text(narratives.get("reflection", "")),
                "story_summary": clean_display_text(case.get("story_summary")),
                "study_method": clean_display_text(case.get("study_method")),
                "turning_point": clean_display_text(case.get("turning_point")),
            },
        }
        desensitized_records.append({
            **common,
            "privacy_version": "desensitized_internal",
            **internal_fields,
        })
        public_sections = [
            {"key": key, "label": label, "text": clean_party_visible_text(narratives.get(key, ""))}
            for key, label in PUBLIC_PARTY_FIELDS
            if clean_party_visible_text(narratives.get(key, ""))
        ]
        public_records.append({
            **common,
            "privacy_version": "public_user",
            "display_name": clean_party_visible_text(case.get("display_name")),
            "age": clean_party_visible_text(display_age(case.get("age_concrete") or case.get("age_band"))),
            "region": clean_party_visible_text(case.get("region")),
            "system": clean_party_visible_text(case.get("system_chip") or case.get("system")),
            "chosen_target": clean_party_visible_text(target),
            "outcome": clean_party_visible_text(case.get("outcome")),
            "tags": compact_tags(case.get("tags") or [], case.get("reason_tags") or [], case.get("story_tags") or [], hide_party=True),
            "sections": public_sections,
        })

    key = lambda item: tuple(item["sort_tuple"])
    return sorted(raw_records, key=key), sorted(desensitized_records, key=key), sorted(public_records, key=key)


def load_exam_rows() -> list[dict[str, str]]:
    workbook = load_workbook(EXAM_SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [clean(item) for item in rows[0]]
    result: list[dict[str, str]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = [clean(item) for item in row]
        if not any(values):
            continue
        result.append({
            "_source_row": str(row_number),
            **{headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))},
        })
    return result


def build_exam_records() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    raw_rows = load_exam_rows()
    raw_records = []
    desensitized_records = []
    public_records = []

    for index, row in enumerate(raw_rows, start=1):
        job = sanitize_exam_private_text(row.get("工作职能"))
        motivation = sanitize_exam_private_text(row.get("考研动机"))
        experience = sanitize_exam_private_text(row.get("备考经验"))
        intent_target, admitted_school, program, precision = exam_intent_and_admitted(row)
        system = infer_exam_system(job, motivation)
        baseline = "；".join(part for part in [row.get("数学基础", ""), row.get("英语基础", "")] if clean(part))
        score_text = f"测评 {row.get('测试总分', '')} 分 → 目标 {row.get('目标总分', '')} 分"
        study_time = "；".join(part for part in [row.get("每周学习时长", ""), row.get("日均学习时长", "")] if clean(part))
        risk = exam_risk(row.get("数学基础", ""), row.get("英语基础", ""), row.get("测试总分", ""), program)
        advice = exam_advice(program, precision, risk)
        richness = len(admitted_school) * 8 + len(intent_target) * 6 + len(motivation) * 3 + len(experience) * 2 + len(baseline) + len(study_time) + (120 if precision == "precise" else 0)
        common = {
            "direction": "exam",
            "source_dataset": str(EXAM_SOURCE.relative_to(ROOT)),
            "source_id": f"exam-row-{row['_source_row']}",
            "display_order_source": index,
            "sort_tuple": [0, 0 if precision == "precise" else 1, -richness, index],
            "richness_score": richness,
            "target_precision": precision,
            "ai_fill_level": "none",
            "ai_filled_fields": [],
            "ai_fill_note": "未使用 AI 补充；风险和建议为规则派生字段。",
            "derived_fields": ["risk", "advice", "system", "goal_tags"],
            "human_review_status": "pending",
        }
        raw_records.append({
            **common,
            "privacy_version": "raw_internal",
            "intent_school": intent_target,
            "admitted_school": admitted_school,
            "source_payload": row,
        })
        internal = {
            "display_name": display_alias(row.get("姓名", ""), row.get("性别", "")),
            "age": age_band(row.get("年龄", "")),
            "system": system,
            "position": job,
            "program": program,
            "intent_school": intent_target,
            "admitted_school": admitted_school,
            "outcome": "上岸",
            "exam_experience": experience,
            "motivation": motivation,
            "baseline": baseline,
            "score": score_text,
            "study_time": study_time,
            "risk": risk,
            "advice": advice,
            "goal_tags": goal_tags(motivation),
            "tags": compact_tags([system, program, "上岸"], goal_tags(motivation)),
        }
        desensitized_records.append({
            **common,
            "privacy_version": "desensitized_internal",
            **internal,
            "hidden_raw_fields": ["姓名", "电话", "毕业院校", "所学专业", "班型", "入班时间"],
        })
        public_records.append({
            **common,
            "privacy_version": "public_user",
            "display_name": internal["display_name"],
            "age": internal["age"],
            "system": internal["system"],
            "program": internal["program"],
            "intent_school": internal["intent_school"],
            "admitted_school": internal["admitted_school"],
            "outcome": internal["outcome"],
            "exam_experience": internal["exam_experience"],
            "motivation": internal["motivation"],
            "baseline": internal["baseline"],
            "score": internal["score"],
            "study_time": internal["study_time"],
            "risk": internal["risk"],
            "advice": internal["advice"],
            "tags": internal["tags"],
        })

    key = lambda item: tuple(item["sort_tuple"])
    return sorted(raw_records, key=key), sorted(desensitized_records, key=key), sorted(public_records, key=key)


def wrap_dataset(name: str, direction: str, privacy_version: str, records: list[dict[str, Any]]) -> dict[str, Any]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "meta": {
            "name": name,
            "direction": direction,
            "privacy_version": privacy_version,
            "generated_at": now,
            "record_count": len(records),
            "review_status": "pending_final_review",
            "sort_rule": "未使用 AI 补充优先；同组内按明确度、丰富度、原始顺序排序。",
            "miniapp_connection": "connected_to_pages_cases_v2",
        },
        "records": records,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def public_miniapp_payload(party_public: list[dict[str, Any]], exam_public: list[dict[str, Any]]) -> dict[str, Any]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    miniapp_party_public = [
        record for record in party_public
        if record.get("ai_fill_level", "none") == "none"
    ]
    return {
        "meta": {
            "name": "案例 V2 小程序公开发布数据",
            "generated_at": now,
            "privacy_version": "public_user",
            "record_count": len(miniapp_party_public) + len(exam_public),
            "party_count": len(miniapp_party_public),
            "exam_count": len(exam_public),
            "party_filter": "only ai_fill_level=none records are included for party direction",
            "exam_filter": "all public exam records are included",
            "source_files": [
                str((DB_DIR / "party-public.json").relative_to(ROOT)),
                str((DB_DIR / "exam-public.json").relative_to(ROOT)),
            ],
            "privacy_note": "仅包含用户端可展示字段；党校方向仅进入未经过 AI 补充的数据；未脱敏和内部脱敏字段不进入小程序包。",
        },
        "records": miniapp_party_public + exam_public,
    }


def render_analytics_dashboard(public_payload: dict[str, Any]) -> str:
    meta = public_payload["meta"]
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    events = [
        {
            "event": "page_view",
            "name": "进入案例 V2 页",
            "questions": "多少人进入案例页、来自哪个入口、默认看到哪个方向。",
            "fields": "page, route, scene, source, page_version, active_tab, party_count, management_exam_count",
        },
        {
            "event": "filter_case_list",
            "name": "切换党校/管综",
            "questions": "用户更关注党校还是管综，切换后是否继续看。",
            "fields": "filter_key=case_type, filter_value, from_tab, current_total_count, current_visible_count",
        },
        {
            "event": "case_card_click",
            "name": "点击案例卡片",
            "questions": "哪些案例最吸引用户，党校/管综哪类卡片点击率更高。",
            "fields": "target_id, target_name, case_type, case_quality, chosen_target, source_id",
        },
        {
            "event": "modal_open",
            "name": "打开案例详情",
            "questions": "用户是否真的进入完整内容阅读。",
            "fields": "modal_type=case_detail, target_id, target_name, case_type, active_tab",
        },
        {
            "event": "modal_close",
            "name": "关闭案例详情",
            "questions": "详情页关闭方式，是否读完后点知道了。",
            "fields": "modal_type=case_detail, close_method, target_id, case_type",
        },
        {
            "event": "case_list_load_more",
            "name": "继续加载更多",
            "questions": "用户是否有继续浏览意愿，列表深度到多少。",
            "fields": "previous_visible_count, next_visible_count, total_count, active_tab",
        },
        {
            "event": "cta_click",
            "name": "案例页转化点击",
            "questions": "是否从案例页去测评、咨询、服务分布等转化入口。",
            "fields": "position, label, route, active_tab, target_id",
        },
    ]
    event_rows = "\n".join(
        f"<tr><td><code>{escape(item['event'])}</code></td><td>{escape(item['name'])}</td><td>{escape(item['questions'])}</td><td>{escape(item['fields'])}</td></tr>"
        for item in events
    )
    kpi_cards = [
        ("曝光", "案例页 PV / UV", "page_view 去重后统计"),
        ("兴趣", "案例卡点击率", "case_card_click / page_view"),
        ("深读", "详情打开率", "modal_open / case_card_click"),
        ("方向", "党校/管综关注占比", "filter_case_list + active_tab"),
        ("深度", "继续加载率", "case_list_load_more / page_view"),
        ("转化", "测评/咨询点击率", "cta_click 按 position 聚合"),
    ]
    kpi_html = "\n".join(
        f"<section class='kpi'><b>{escape(title)}</b><strong>{escape(metric)}</strong><span>{escape(desc)}</span></section>"
        for title, metric, desc in kpi_cards
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>用户行为看板说明</title>
  <style>
    :root {{ --bg:#f7f0e6; --paper:#fffdfa; --line:#e6d8c8; --text:#2a251e; --muted:#766b60; --accent:#c96f3a; --green:#2f7a4c; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; background:var(--bg); color:var(--text); }}
    header {{ background:linear-gradient(135deg,#fffdfa,#fff3e8); border-bottom:1px solid var(--line); }}
    .wrap {{ max-width:1180px; margin:0 auto; padding:24px; }}
    h1 {{ margin:0; font-size:28px; }}
    .sub {{ margin-top:8px; color:var(--muted); line-height:1.7; }}
    .grid {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:12px; margin:18px 0; }}
    .card,.kpi,table {{ background:var(--paper); border:1px solid var(--line); border-radius:18px; box-shadow:0 8px 24px rgba(60,50,40,.055); }}
    .card {{ padding:18px; }}
    .card b,.kpi b {{ display:block; color:var(--accent); font-size:13px; margin-bottom:8px; }}
    .card strong {{ display:block; font-size:30px; }}
    .card span,.kpi span {{ color:var(--muted); font-size:13px; line-height:1.6; }}
    .kpi {{ padding:16px; }}
    .kpi strong {{ display:block; font-size:18px; margin-bottom:6px; }}
    h2 {{ margin:26px 0 12px; font-size:20px; }}
    table {{ width:100%; border-collapse:collapse; overflow:hidden; }}
    th,td {{ border-bottom:1px solid var(--line); padding:12px; text-align:left; vertical-align:top; font-size:13px; line-height:1.65; }}
    th {{ color:var(--muted); background:#fbf7f1; }}
    code {{ color:#8a4d2d; background:#fff5ec; padding:2px 5px; border-radius:6px; }}
    .note {{ padding:16px; border-left:4px solid var(--accent); background:#fffdfa; border-radius:14px; color:var(--muted); line-height:1.8; }}
    @media (max-width: 860px) {{ .grid {{ grid-template-columns:1fr; }} .wrap {{ padding:18px; }} }}
  </style>
</head>
<body>
  <header>
    <div class="wrap">
      <h1>用户行为看板说明</h1>
      <div class="sub">生成时间：{escape(generated_at)} · 本页说明上线后能通过埋点看到哪些数据，不包含真实用户明细。当前小程序公开案例：党校 {meta['party_count']} 条，管综 {meta['exam_count']} 条。</div>
    </div>
  </header>
  <main class="wrap">
    <div class="grid">
      <section class="card"><b>公开案例总量</b><strong>{meta['record_count']}</strong><span>小程序只接入用户端公开字段</span></section>
      <section class="card"><b>党校案例</b><strong>{meta['party_count']}</strong><span>可看 Tab 点击、卡片点击、详情打开</span></section>
      <section class="card"><b>管综案例</b><strong>{meta['exam_count']}</strong><span>可看院校、备考经验、基础字段关注度</span></section>
    </div>

    <h2>核心看板指标</h2>
    <div class="grid">{kpi_html}</div>

    <h2>事件字典</h2>
    <table>
      <thead><tr><th>事件</th><th>含义</th><th>回答的问题</th><th>关键字段</th></tr></thead>
      <tbody>{event_rows}</tbody>
    </table>

    <h2>使用口径</h2>
    <div class="note">上线后，后台只要按 <code>page=cases_v2</code> 过滤即可看案例 V2 页面行为。重点关注：进入量、党校/管综切换占比、卡片点击排行、详情打开率、加载更多率、从案例页去测评/咨询的点击率。</div>
  </main>
</body>
</html>
"""


def render_record_card(record: dict[str, Any], privacy_version: str) -> str:
    ai = record.get("ai_fill_level", "none")
    ai_label = "未使用 AI 补充" if ai == "none" else f"AI 补充：{', '.join(record.get('ai_filled_fields') or [])}"
    ai_class = "ai-none" if ai == "none" else "ai-partial"
    if record["direction"] == "party":
        if privacy_version == "raw_internal":
            source = record["source_payload"]
            body = "".join(
                f"<div><span>{escape(key)}</span>{escape(visible_source_value(value))}</div>"
                for key, value in source.items()
                if has_visible_source_value(visible_source_value(value))
            )
            title = f"{source.get('display_name', '')} · {display_age(source.get('age_concrete') or source.get('age_band') or '')} · {source.get('region', '')}"
        elif privacy_version == "desensitized_internal":
            fields = record.get("fields", {})
            body = "".join(f"<div><span>{escape(key)}</span>{escape(value)}</div>" for key, value in fields.items() if clean(value))
            title = f"{record.get('display_name', '')} · {record.get('age', '')} · {record.get('region', '')}"
        else:
            body = "".join(f"<div><span>{escape(section['label'])}</span>{escape(section['text'])}</div>" for section in record.get("sections", []))
            title = f"{record.get('display_name', '')} · {record.get('age', '')}{' · ' + record.get('region', '') if record.get('region') else ''}"
        target = record.get("chosen_target") or record.get("source_payload", {}).get("chosen_school", "")
        tags = record.get("tags") or compact_tags(record.get("source_payload", {}).get("tags", []))
    else:
        if privacy_version == "raw_internal":
            source = record["source_payload"]
            title = f"{source.get('姓名', '')} · {source.get('年龄', '')} · {source.get('上岸院校', '')}"
            body = (
                f"<div><span>意向院校</span>{escape(record.get('intent_school', ''))}</div>"
                f"<div><span>上岸院校</span>{escape(record.get('admitted_school', ''))}</div>"
            )
            body += "".join(f"<div><span>{escape(key)}</span>{escape(value)}</div>" for key, value in source.items() if clean(value))
            target = source.get("上岸院校", "")
            tags = []
        else:
            title = f"{record.get('display_name', '')} · {record.get('age', '')}"
            target = record.get("admitted_school", "")
            body_keys = [
                ("意向院校", "intent_school"),
                ("上岸院校", "admitted_school"),
                ("备考经验", "exam_experience"),
                ("考研动机", "motivation"),
                ("基础", "baseline"),
                ("测评/目标", "score"),
                ("备考投入", "study_time"),
            ]
            if privacy_version == "desensitized_internal":
                body_keys.extend([("阻力评估", "risk"), ("备考建议", "advice")])
            body = "".join(f"<div><span>{label}</span>{escape(record.get(key, ''))}</div>" for label, key in body_keys if clean(record.get(key, "")))
            tags = record.get("tags", [])

    tag_html = "".join(f'<span class="tag">{escape(tag)}</span>' for tag in tags if clean(tag))
    return f"""
    <article class="case-card" data-direction="{escape(record['direction'])}" data-version="{escape(privacy_version)}">
      <div class="card-head">
        <div>
          <div class="rank">{escape(record['source_id'])} · {escape(privacy_version)}</div>
          <h3>{escape(title)}</h3>
        </div>
        <div class="score">{escape(record.get('richness_score', ''))}</div>
      </div>
      <div class="ai-line"><span class="ai-badge {ai_class}">{escape(ai_label)}</span></div>
      <div class="target">目标/结果：<strong>{escape(target)}</strong></div>
      <div class="tags">{tag_html}</div>
      <div class="fact-grid">{body}</div>
    </article>
    """


def render_html(datasets: dict[str, dict[str, Any]]) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cards = []
    for key, dataset in datasets.items():
        version = dataset["meta"]["privacy_version"]
        for record in dataset["records"]:
            cards.append(render_record_card(record, version))
    summary_rows = "".join(
        f"<tr><td>{escape(data['meta']['direction'])}</td><td>{escape(data['meta']['privacy_version'])}</td><td>{data['meta']['record_count']}</td><td>{escape(key)}</td></tr>"
        for key, data in datasets.items()
    )
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>案例 V2 入库核对页</title>
  <style>
    :root {{ --bg:#f7f0e6; --paper:#fffdfa; --line:#e6d8c8; --text:#2a251e; --muted:#766b60; --accent:#c96f3a; --green:#2f7a4c; --blue:#355f8f; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; background:var(--bg); color:var(--text); }}
    header {{ position:sticky; top:0; z-index:10; background:rgba(255,253,250,.94); border-bottom:1px solid var(--line); backdrop-filter:blur(12px); }}
    .bar {{ max-width:1260px; margin:0 auto; padding:18px 22px; display:flex; justify-content:space-between; gap:18px; align-items:center; }}
    h1 {{ margin:0; font-size:24px; }}
    .subtitle {{ margin-top:6px; color:var(--muted); font-size:13px; }}
    .tools {{ display:grid; gap:8px; justify-items:end; }}
    .tool-row {{ display:flex; gap:8px; flex-wrap:wrap; justify-content:flex-end; }}
    .tool-label {{ align-self:center; color:var(--muted); font-size:12px; font-weight:800; letter-spacing:.04em; }}
    button {{ border:1px solid var(--line); border-radius:999px; background:white; color:var(--muted); padding:10px 14px; font-weight:800; cursor:pointer; }}
    button.active {{ background:var(--accent); color:white; border-color:var(--accent); }}
    main {{ max-width:1260px; margin:0 auto; padding:22px; }}
    .note, table {{ background:var(--paper); border:1px solid var(--line); border-radius:16px; padding:14px; margin-bottom:16px; }}
    table {{ width:100%; border-collapse:collapse; overflow:hidden; }}
    td, th {{ border-bottom:1px solid var(--line); padding:9px; text-align:left; font-size:13px; }}
    .case-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:14px; }}
    .case-card {{ background:white; border:1px solid var(--line); border-radius:20px; padding:18px; box-shadow:0 8px 24px rgba(60,50,40,.055); }}
    .card-head {{ display:flex; justify-content:space-between; gap:12px; align-items:flex-start; }}
    .rank {{ color:var(--accent); font-size:12px; font-weight:800; }}
    h3 {{ margin:5px 0 0; font-size:18px; line-height:1.45; }}
    .score {{ min-width:56px; text-align:center; padding:8px 10px; border-radius:14px; background:#fff5ec; color:var(--accent); font-weight:800; }}
    .ai-line {{ margin-top:8px; }}
    .ai-badge {{ display:inline-flex; border-radius:999px; padding:5px 9px; font-size:12px; line-height:1; font-weight:800; }}
    .ai-none {{ color:var(--green); background:#eef7f0; }}
    .ai-partial {{ color:#8a4d2d; background:#fff5ec; }}
    .target {{ margin-top:12px; line-height:1.6; }}
    .target strong {{ color:var(--accent); }}
    .tags {{ display:flex; gap:6px; flex-wrap:wrap; margin-top:10px; }}
    .tag {{ display:inline-flex; padding:5px 9px; border-radius:999px; background:#f1e8dc; color:#5b4c40; font-size:12px; line-height:1; }}
    .fact-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:8px; margin-top:12px; }}
    .fact-grid div {{ background:#fbf7f1; border-radius:12px; padding:10px; color:#4c433a; font-size:13px; line-height:1.65; }}
    .fact-grid span {{ display:block; color:var(--muted); font-size:12px; font-weight:800; margin-bottom:3px; }}
    .case-card[data-direction="party"][data-version="public_user"] .fact-grid {{ grid-template-columns:1fr; gap:12px; }}
    .case-card[data-direction="exam"][data-version="public_user"] .fact-grid {{ grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:10px; }}
    .case-card[data-version="public_user"] .fact-grid div {{ background:#fffdfa; border:1px solid var(--line); border-left:4px solid var(--accent); border-radius:14px; padding:13px 14px; font-size:14px; line-height:1.8; }}
    .case-card[data-version="public_user"] .fact-grid span {{ color:var(--accent); font-size:13px; margin-bottom:5px; }}
    .hidden {{ display:none !important; }}
    @media (max-width: 860px) {{ .bar {{ display:block; }} .tools {{ margin-top:14px; justify-items:start; }} .tool-row {{ justify-content:flex-start; }} .case-grid,.fact-grid {{ grid-template-columns:1fr; }} }}
  </style>
</head>
<body>
  <header>
    <div class="bar">
      <div>
        <h1>案例 V2 入库核对页</h1>
        <div class="subtitle">生成时间：{escape(generated_at)} · 三版本切换核对 · 已接入小程序 V2，老案例页保留</div>
      </div>
      <div class="tools">
        <div class="tool-row">
          <span class="tool-label">方向</span>
          <button class="active" data-direction="party">党校</button>
          <button data-direction="exam">管综</button>
        </div>
        <div class="tool-row">
          <span class="tool-label">版本</span>
          <button class="active" data-version="public_user">用户端展示版</button>
          <button data-version="desensitized_internal">脱敏内部版</button>
          <button data-version="raw_internal">未脱敏内部版</button>
        </div>
      </div>
    </div>
  </header>
  <main>
    <div class="note">说明：未脱敏内部版仅供内部核对，不进入小程序包；用户端展示版隐藏了部分字段；所有版本均标注 AI 补充情况，并优先显示未使用 AI 补充的数据。</div>
    <table>
      <thead><tr><th>方向</th><th>版本</th><th>条数</th><th>文件</th></tr></thead>
      <tbody>{summary_rows}</tbody>
    </table>
    <div class="case-grid">{''.join(cards)}</div>
  </main>
  <script>
    let activeDirection = 'party';
    let activeVersion = 'public_user';
    const buttons = Array.from(document.querySelectorAll('button'));
    const cards = Array.from(document.querySelectorAll('.case-card'));
    function applyFilter() {{
      cards.forEach(card => {{
        const visible = card.dataset.direction === activeDirection && card.dataset.version === activeVersion;
        card.classList.toggle('hidden', !visible);
      }});
      buttons.forEach(btn => {{
        if (btn.dataset.direction) btn.classList.toggle('active', btn.dataset.direction === activeDirection);
        if (btn.dataset.version) btn.classList.toggle('active', btn.dataset.version === activeVersion);
      }});
    }}
    buttons.forEach(btn => btn.addEventListener('click', () => {{
      if (btn.dataset.direction) activeDirection = btn.dataset.direction;
      if (btn.dataset.version) activeVersion = btn.dataset.version;
      applyFilter();
    }}));
    applyFilter();
  </script>
</body>
</html>
"""


def main() -> None:
    party_raw, party_desensitized, party_public = build_party_records()
    exam_raw, exam_desensitized, exam_public = build_exam_records()
    datasets = {
        "party-raw-internal.json": wrap_dataset("党校未脱敏内部版", "party", "raw_internal", party_raw),
        "party-desensitized-internal.json": wrap_dataset("党校脱敏内部版", "party", "desensitized_internal", party_desensitized),
        "party-public.json": wrap_dataset("党校用户端展示版", "party", "public_user", party_public),
        "exam-raw-internal.json": wrap_dataset("管综未脱敏内部版", "exam", "raw_internal", exam_raw),
        "exam-desensitized-internal.json": wrap_dataset("管综脱敏内部版", "exam", "desensitized_internal", exam_desensitized),
        "exam-public.json": wrap_dataset("管综用户端展示版", "exam", "public_user", exam_public),
    }
    for filename, payload in datasets.items():
        write_json(DB_DIR / filename, payload)
    miniapp_payload = public_miniapp_payload(party_public, exam_public)
    write_json(MINIAPP_PUBLIC_DATA, miniapp_payload)
    write_json(DB_DIR / "manifest.json", {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "miniapp_connection": "connected_to_pages_cases_v2",
        "files": [
            {"file": filename, **payload["meta"]}
            for filename, payload in datasets.items()
        ],
    })
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(render_html(datasets), encoding="utf-8")
    ANALYTICS_DASHBOARD.parent.mkdir(parents=True, exist_ok=True)
    ANALYTICS_DASHBOARD.write_text(render_analytics_dashboard(miniapp_payload), encoding="utf-8")
    print(f"wrote {DB_DIR.relative_to(ROOT)}")
    print(f"wrote {REPORT.relative_to(ROOT)}")
    print(f"wrote {MINIAPP_PUBLIC_DATA.relative_to(ROOT)}")
    print(f"wrote {ANALYTICS_DASHBOARD.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
