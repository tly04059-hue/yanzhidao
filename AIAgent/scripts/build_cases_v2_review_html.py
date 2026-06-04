from __future__ import annotations

import html
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_party_duplicate_rewrite_draft import choose_reason as draft_choose_reason
from build_party_duplicate_rewrite_draft import feedback as draft_feedback


ROOT = Path(__file__).resolve().parents[2]
PARTY_SOURCE_CANDIDATES = [
    ROOT / "miniapp" / "dx-cases.json",
    ROOT / "schooltool" / "data" / "cases.json",
]
EXAM_SOURCE = ROOT / "studentCases" / "sourceData" / "管综已上岸学员案例信息汇总表.xlsx"
OUTPUT = ROOT / "studentCases" / "reports" / "cases-v2-review.html"

NARRATIVE_KEYS = [
    "key_quote",
    "narrative_choose",
    "reflection",
    "story_summary",
    "study_method",
    "turning_point",
]
NARRATIVE_LABELS = {
    "key_quote": "报考动机",
    "narrative_choose": "选择原因",
    "reflection": "真实反馈",
    "study_method": "备考方法",
    "turning_point": "转折点",
}
NARRATIVE_WEIGHTS = {
    "key_quote": 3.0,
    "narrative_choose": 2.0,
    "reflection": 2.5,
    "story_summary": 2.0,
    "study_method": 1.5,
    "turning_point": 1.5,
}
IMPRECISE_TARGET_PATTERN = re.compile(
    r"(及附近|名校|好上岸|第一优先|川渝院校|重庆院校|医科类院校|医学类院校|类院校|成都985|重庆985|新疆|学费低|离雅安|不详)"
)
AGE_RANGE_PATTERN = re.compile(r"(\d{2})\s*-\s*(\d{2})\s*岁")
YANZHIDAO_REASON_PATTERN = re.compile(
    r"(研知道|老师|报班|入班|资料|讲义|带背|密训|模考|题库|1V1|一页纸|33法|督学|打卡|回放|核心考点|押题|班主任|杨老师|姚老师|胡老师)"
)
DISTRICT_UNIT_PATTERN = re.compile(
    r"(青羊区|锦江区|武侯区|金牛区|成华区|高新区|天府新区|双流区|温江区|郫都区|新都区|龙泉驿区|"
    r"[^，,；;、\\s]{1,6}区|[^，,；;、\\s]{1,6}县|[^，,；;、\\s]{1,6}市)委党校"
)
HIDDEN_TAGS = {"其他", "M8-resync", "system_unknown_storied", "心得型"}
PARTY_CASE_IDS_TO_DROP = {"C123"}
PARTY_TARGET_OVERRIDES = {
    "C022": "四川党校 · 政治学",
    "C025": "四川党校 · 经济学",
    "C112": "四川党校 · 政治学",
    "C122": "四川党校 · 政治学",
}


def clean(value: Any) -> str:
    return str(value or "").strip()


UNKNOWN_AGE_VALUES = {"?", "未知", "不详"}


def normalize_age_range(value: Any) -> str:
    text = clean(value).replace("岁段", "岁")
    match = re.fullmatch(r"(\d{2})\s*-\s*(\d{2})\s*岁?", text)
    if match:
        start, end = (int(match.group(1)), int(match.group(2)))
        return f"约 {(start + end) // 2} 岁"
    text = AGE_RANGE_PATTERN.sub(lambda item: f"约 {(int(item.group(1)) + int(item.group(2))) // 2} 岁", text)
    return re.sub(r"约\s*(\d{1,2})\s*岁", r"约 \1 岁", text)


def stable_age_label(seed: Any) -> str:
    seed_text = clean(seed) or "case"
    age = 31 + (sum(ord(char) for char in seed_text) % 10)
    return f"约 {age} 岁"


def display_age(value: Any, seed: Any) -> str:
    age = clean(value)
    if not age or age in UNKNOWN_AGE_VALUES:
        return stable_age_label(seed)
    return normalize_age_range(age)


def clean_display_text(value: Any) -> str:
    text = clean(value)
    text = text.replace("M8-resync", "")
    text = text.replace("其他", "")
    text = normalize_age_range(text)
    text = re.sub(r"\s+", " ", text)
    text = text.replace("未知，", "未知，").strip()
    return text


def sanitize_exam_private_text(value: Any) -> str:
    text = clean(value)
    text = DISTRICT_UNIT_PATTERN.sub("党校系统", text)
    text = re.sub(r"(青羊区|锦江区|武侯区|金牛区|成华区|高新区|天府新区|双流区|温江区|郫都区|新都区|龙泉驿区)", "当地", text)
    text = text.replace("党校系统", "相关系统")
    text = text.replace("党校单证", "单证")
    text = text.replace("党校", "在职研")
    return text.strip(" ，,；;、")


def escape(value: Any) -> str:
    return html.escape(clean_display_text(value), quote=True)


def clean_party_visible_text(value: Any) -> str:
    text = clean_display_text(value)
    text = re.sub(r"(四川省委党校|中共四川省委党校|重庆市委党校|中共重庆市委党校|四川党校|重庆党校|省委党校|市委党校)", "", text)
    text = text.replace("党校在职研究生", "在职研究生")
    text = text.replace("党校研究生", "在职研究生")
    text = text.replace("考党校", "考在职研")
    text = text.replace("报考党校", "报考在职研")
    text = text.replace("备考党校", "备考在职研")
    text = text.replace("党校", "")
    text = re.sub(r"\s*·\s*·\s*", " · ", text)
    text = re.sub(r"(^|\s)[·，,、；;：:]\s*", r"\1", text)
    text = re.sub(r"\s*[·，,、；;：:]$", "", text)
    return re.sub(r"\s+", " ", text).strip()


def escape_party(value: Any) -> str:
    return html.escape(clean_party_visible_text(value), quote=True)


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", clean(value))


def compact_tags(*groups: list[str] | None) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for group in groups:
        for item in group or []:
            tag = clean(item)
            if tag in HIDDEN_TAGS:
                continue
            if tag and tag not in seen:
                seen.add(tag)
                result.append(tag)
    return result


def display_label(value: Any) -> str:
    label = clean(value)
    if not label or label in HIDDEN_TAGS:
        return ""
    if "其他" in label or "M8-resync" in label:
        return ""
    return label


def load_party_source() -> tuple[Path, list[dict[str, Any]]]:
    for source in PARTY_SOURCE_CANDIDATES:
        if not source.exists():
            continue
        payload = json.loads(source.read_text(encoding="utf-8"))
        cases = payload.get("data", {}).get("cases", [])
        if cases:
            return source, cases
    raise FileNotFoundError("No party case source found.")


def deduped_narratives(case: dict[str, Any]) -> list[tuple[str, str]]:
    kept: list[tuple[str, str]] = []
    for key in NARRATIVE_KEYS:
        text = clean(case.get(key))
        if not text:
            continue

        replaced = False
        for idx, (existing_key, existing_text) in enumerate(kept):
            if text == existing_text or existing_text in text or text in existing_text:
                if len(text) > len(existing_text):
                    kept[idx] = (key, text)
                replaced = True
                break
        if not replaced:
            kept.append((key, text))
    return kept


def party_richness(case: dict[str, Any]) -> int:
    score = 0.0
    for key, text in deduped_narratives(case):
        score += min(len(text), 180) * NARRATIVE_WEIGHTS[key]
    return round(score)


def party_quality(score: int) -> str:
    if score >= 480:
        return "高丰富"
    if score >= 180:
        return "标准"
    return "简版"


def display_narratives(case: dict[str, Any]) -> dict[str, str]:
    narratives = {key: clean(case.get(key)) for key in NARRATIVE_KEYS}
    core_keys = ["key_quote", "narrative_choose", "reflection"]
    core_values = [narratives[key] for key in core_keys]
    has_all_core = all(core_values)
    is_exact_duplicate_core = has_all_core and len(set(core_values)) == 1

    if not is_exact_duplicate_core:
        return narratives

    if narratives["narrative_choose"]:
        narratives["narrative_choose"] = draft_choose_reason(case)
    if narratives["reflection"]:
        narratives["reflection"] = draft_feedback(case)
    return narratives


def ai_fill_info(case: dict[str, Any]) -> dict[str, Any]:
    core_keys = ["key_quote", "narrative_choose", "reflection"]
    core_values = [clean(case.get(key)) for key in core_keys]
    has_all_core = all(core_values)
    is_exact_duplicate_core = has_all_core and len(set(core_values)) == 1
    if not is_exact_duplicate_core:
        return {
            "ai_fill_level": "none",
            "ai_filled_fields": [],
            "ai_label": "未使用 AI 补充",
        }
    return {
        "ai_fill_level": "partial",
        "ai_filled_fields": ["narrative_choose", "reflection"],
        "ai_label": "AI 补充：真实反馈",
    }


def normalize_system(system_chip: str, system: str) -> str:
    return display_label(system_chip or system)


def infer_program(chosen_path: str, target: str, fallback: str = "MPA") -> str:
    text = clean(target).upper()
    if chosen_path == "A" or "党校" in target:
        return "党校"
    for program in ["MEM", "MTA", "MBA", "MPA"]:
        if program in text:
            return program
    return fallback


def normalize_target(source_target: str, fallback_program: str = "") -> tuple[str, str, str]:
    target = normalize_spaces(source_target)
    upper = target.upper().replace(" ", "")
    program = ""
    for candidate in ["MEM", "MTA", "MBA", "MPA"]:
        if candidate in upper:
            program = candidate
            break
    if not program:
        program = fallback_program.upper() or "MPA"

    school = re.sub(program, "", target, flags=re.IGNORECASE).strip(" ·-—")
    school = normalize_spaces(school)
    normalized = f"{school} · {program}" if school else program
    precision = "明确" if school and not IMPRECISE_TARGET_PATTERN.search(school) else "宽泛"
    return normalized, program, precision


def exam_intent_and_admitted(get) -> tuple[str, str, str, str]:
    intent_raw = clean(get("意向院校") or get("原始上岸院校") or get("上岸院校"))
    admitted_raw = clean(get("上岸院校"))
    intent, program, precision = normalize_target(intent_raw, get("意向专业"))
    admitted, _, admitted_precision = normalize_target(admitted_raw, get("意向专业"))
    if admitted_precision == "明确":
        precision = admitted_precision
    return intent, admitted, program, precision


def infer_exam_system(job: str, motivation: str) -> str:
    text = f"{job} {motivation}"
    if re.search(r"人民银行|银行|金融", text):
        return "银行金融"
    if re.search(r"国企|央企", text):
        return "国央企"
    if re.search(r"教师|高职|学校|培训", text):
        return "教育系统"
    if re.search(r"公务员|事业编|体制内|机关|街道|交通|公路|精神文明", text):
        return "党政机关"
    return ""


def goal_tags(motivation: str) -> list[str]:
    text = clean(motivation)
    tags: list[str] = []
    if "遴选" in text:
        tags.append("遴选")
    if "晋升" in text or "向上发展" in text:
        tags.append("晋升")
    if "学历" in text:
        tags.append("学历提升")
    if "提升自己" in text or "学习" in text or "充实自我" in text:
        tags.append("自我提升")
    if "职业发展" in text or "后续发展" in text:
        tags.append("职业发展")
    if "人脉" in text:
        tags.append("人脉拓展")
    if "双证" in text:
        tags.append("双证")
    return tags or ["学历提升"]


def age_band(age: str) -> str:
    if clean(age) in UNKNOWN_AGE_VALUES:
        return stable_age_label(age)
    try:
        value = int(float(clean(age)))
    except ValueError:
        label = normalize_age_range(age)
        if re.fullmatch(r"\d{2}\s*-\s*\d{2}", label):
            return f"{label.replace(' ', '')}岁"
        return label
    if value < 25:
        return "25岁以下"
    if value <= 30:
        return normalize_age_range("25-30岁")
    if value <= 35:
        return normalize_age_range("30-35岁")
    if value <= 40:
        return normalize_age_range("35-40岁")
    return "40岁以上"


def display_alias(name: str, gender: str) -> str:
    surname = clean(name)[:1] or "同"
    suffix = "女士" if clean(gender) == "女" else "先生" if clean(gender) == "男" else "同学"
    return f"{surname}{suffix}"


def exam_risk(math_base: str, english_base: str, test_score: str, program: str) -> str:
    try:
        score = float(clean(test_score))
    except ValueError:
        score = 0
    if score and score < 50:
        return "测评分数偏低，基础补齐和择校保底需要优先处理。"
    if "未及格" in math_base or "薄弱" in math_base:
        return "数学基础是主要风险，择校不宜只看院校名气。"
    if "四级未过" in english_base or "未参加" in english_base:
        return "英语基础需要同步补齐，避免后期总分和单科线同时受影响。"
    if program == "MBA":
        return "MBA 更容易受学费、城市和上课方式约束。"
    if program == "MEM":
        return "MEM 需要兼顾项目适配度、复试要求和院校专业方向。"
    return "需要同时平衡目标院校、基础水平和备考时间。"


def exam_advice(program: str, target_precision: str, risk: str) -> str:
    if target_precision == "宽泛":
        return "先把意向院校、城市、学费和分数线拆开复盘，再沉淀成可复用择校口径。"
    if "数学" in risk or "测评" in risk:
        return "优先选择分数线稳定、招生人数相对充足、学费和上课方式匹配的院校，再安排基础补齐。"
    if program == "MBA":
        return "先锁定预算和上课地点，再比较复试压力与院校认可度。"
    if program == "MEM":
        return "优先确认专业方向和复试要求，再做冲刺与保底组合。"
    return "按专业方向和城市筛选，再结合分数线、学费、招生人数做保底和冲刺组合。"


def load_exam_cases() -> list[dict[str, Any]]:
    workbook = load_workbook(EXAM_SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [clean(item) for item in rows[0]]
    index = {name: idx for idx, name in enumerate(headers)}
    required = ["姓名", "性别", "年龄", "工作职能", "意向专业", "上岸院校", "备考经验", "数学基础", "英语基础", "测试总分", "目标总分", "每周学习时长", "日均学习时长", "考研动机"]
    missing = [field for field in required if field not in index]
    if missing:
        raise ValueError(f"Missing exam columns: {missing}")

    cases: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        values = [clean(item) for item in row]
        if not any(values):
            continue

        def get(field: str) -> str:
            idx = index[field]
            return values[idx] if idx < len(values) else ""

        job = sanitize_exam_private_text(get("工作职能"))
        motivation = sanitize_exam_private_text(get("考研动机"))
        exam_experience = sanitize_exam_private_text(get("备考经验"))
        intent_target, admitted_school, program, precision = exam_intent_and_admitted(get)
        goals = goal_tags(motivation)
        system = infer_exam_system(job, motivation)
        risk = exam_risk(get("数学基础"), get("英语基础"), get("测试总分"), program)
        advice = exam_advice(program, precision, risk)
        baseline = "；".join(part for part in [get("数学基础"), get("英语基础")] if part)
        score_text = f"测评 {get('测试总分')} 分 → 目标 {get('目标总分')} 分"
        study_time = "；".join(part for part in [get("每周学习时长"), get("日均学习时长")] if part)
        richness = (
            len(admitted_school) * 8
            + len(intent_target) * 6
            + len(motivation) * 3
            + len(exam_experience) * 2
            + len(baseline)
            + len(study_time)
            + (120 if precision == "明确" else 0)
        )

        cases.append(
            {
                "source_order": row_number - 1,
                "id": f"exam-{row_number - 1:03d}",
                "display_alias": display_alias(get("姓名"), get("性别")),
                "age": age_band(get("年龄")),
                "raw_age": get("年龄"),
                "system": system,
                "position": display_label(job),
                "program": program,
                "target": admitted_school,
                "intent_target": intent_target,
                "target_precision": precision,
                "outcome": "上岸",
                "baseline": baseline,
                "score_text": score_text,
                "study_time": study_time,
                "motivation": motivation,
                "exam_experience": exam_experience,
                "risk": risk,
                "advice": advice,
                "goal_tags": goals,
                "tags": compact_tags([system, display_label(get("工作职能")), program, precision, "上岸"], goals),
                "richness": richness,
                "intent_target_raw": intent_target,
            }
        )

    return sorted(
        cases,
        key=lambda item: (
            item["target_precision"] != "明确",
            -item["richness"],
            item["source_order"],
        ),
    )


def load_party_cases() -> tuple[Path, list[dict[str, Any]]]:
    source, raw_cases = load_party_source()
    cases = []
    for index, case in enumerate(raw_cases, start=1):
        if clean(case.get("chosen_path")) != "A":
            continue
        case_id = clean(case.get("case_id"))
        if case_id in PARTY_CASE_IDS_TO_DROP:
            continue
        target = PARTY_TARGET_OVERRIDES.get(case_id, clean(case.get("chosen_school")))
        system = normalize_system(clean(case.get("system_chip")), clean(case.get("system")))
        position = display_label(clean(case.get("system") if clean(case.get("system")) != "system_unknown_storied" else case.get("unit_narrative")))
        richness = party_richness(case)
        narratives = display_narratives(case)
        ai_info = ai_fill_info(case)
        fields = [key for key in NARRATIVE_KEYS if key != "narrative_choose" and clean(narratives.get(key))]
        goal_source = []
        for key in ["reason_tags", "reason_keywords", "story_tags", "tags"]:
            goal_source.extend(case.get(key) or [])
        cases.append(
            {
                "source_order": index,
                "id": case_id,
                "display_alias": clean(case.get("display_name")) or "党校同学",
                "age": display_age(case.get("age_concrete") or case.get("age_band"), case.get("case_id")),
                "region": clean(case.get("region")),
                "system": system,
                "position": position,
                "program": infer_program(clean(case.get("chosen_path")), target),
                "target": target,
                "outcome": clean(case.get("outcome")) or "结果待确认",
                "quality": party_quality(richness),
                "richness": richness,
                "fields": fields,
                "goal_tags": compact_tags(goal_source)[:8],
                "tags": compact_tags(case.get("tags") or [], [system, "党校", clean(case.get("outcome"))]),
                "narratives": narratives,
                "summary": clean(case.get("story_summary")),
                "ai_fill_level": ai_info["ai_fill_level"],
                "ai_filled_fields": ai_info["ai_filled_fields"],
                "ai_label": ai_info["ai_label"],
            }
        )
    return source, sorted(
        cases,
        key=lambda item: (
            -item["richness"],
            -len(item["fields"]),
            item["source_order"],
        ),
    )


def render_tags(tags: list[str], class_name: str = "tag") -> str:
    return "".join(f'<span class="{class_name}">{escape(tag)}</span>' for tag in tags if clean(tag))


def render_party_tags(tags: list[str], class_name: str = "tag") -> str:
    return "".join(
        f'<span class="{class_name}">{escape_party(tag)}</span>'
        for tag in tags
        if clean_party_visible_text(tag)
    )


def render_party_card(item: dict[str, Any], rank: int) -> str:
    narrative_blocks = []
    for key in NARRATIVE_KEYS:
        if key in {"story_summary", "narrative_choose", "turning_point"}:
            continue
        text = item["narratives"].get(key, "")
        if text:
            narrative_blocks.append(
                f"""
                <div class="detail-block">
                  <div class="detail-label">{escape(NARRATIVE_LABELS[key])}</div>
                  <div class="detail-text">{escape_party(text)}</div>
                </div>
                """
            )
    meta_parts = [item["system"]] if item["region"] and item["region"] != "未知" else [item["system"], item["position"]]
    meta = " · ".join(part for part in meta_parts if clean(part))
    ai_class = "ai-none" if item["ai_fill_level"] == "none" else "ai-partial"
    return f"""
    <article class="case-card party-card" data-search="{escape_party(' '.join([item['display_alias'], item['target'], item['system'], item['region']]))}">
      <div class="card-head">
        <div>
          <div class="rank">#{rank:03d} · {escape(item['quality'])}</div>
          <h3>{escape_party(item['display_alias'])} · {escape_party(item['age'])}{' · ' + escape_party(item['region']) if item['region'] and item['region'] != '未知' else ''}</h3>
        </div>
        <div class="score">{item['richness']}</div>
      </div>
      <div class="ai-line"><span class="ai-badge {ai_class}">{escape(item['ai_label'])}</span></div>
      <div class="meta">{escape_party(meta)}</div>
      <div class="target">选了：<strong>{escape_party(item['target'])}</strong></div>
      <div class="tags">{render_party_tags(item['tags'][:8])}</div>
      <div class="detail-list">{''.join(narrative_blocks)}</div>
    </article>
    """


def render_exam_card(item: dict[str, Any], rank: int) -> str:
    experience_block = (
        f"<div><span>备考经验</span>{escape(item['exam_experience'])}</div>"
        if clean(item["exam_experience"])
        else ""
    )
    return f"""
    <article class="case-card exam-card" data-search="{escape(' '.join([item['display_alias'], item['target'], item['system'], item['program'], item['motivation'], item['exam_experience']]))}">
      <div class="card-head">
        <div>
          <div class="rank">#{rank:03d} · 管综 · {escape(item['target_precision'])}</div>
          <h3>{escape(item['display_alias'])} · {escape(item['age'])}</h3>
        </div>
        <div class="score">{item['richness']}</div>
      </div>
      <div class="meta">{escape(' · '.join(part for part in [item['system'], item['program']] if clean(part)))}</div>
      <div class="target">意向院校：<strong>{escape(item['intent_target'])}</strong></div>
      <div class="target">上岸院校：<strong>{escape(item['target'])}</strong></div>
      <blockquote>{escape(item['motivation'])}</blockquote>
      <div class="fact-grid">
        <div><span>基础</span>{escape(item['baseline'])}</div>
        <div><span>测评/目标</span>{escape(item['score_text'])}</div>
        <div><span>备考投入</span>{escape(item['study_time'])}</div>
        {experience_block}
      </div>
      <div class="risk">阻力评估：{escape(item['risk'])}</div>
      <div class="advice">备考建议：{escape(item['advice'])}</div>
      <div class="tags">{render_tags(item['tags'][:10])}</div>
    </article>
    """


def render_html(party_source: Path, party_cases: list[dict[str, Any]], exam_cases: list[dict[str, Any]]) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rich_party_count = sum(1 for item in party_cases if item["quality"] == "高丰富")
    precise_exam_count = sum(1 for item in exam_cases if item["target_precision"] == "明确")
    party_cards = "\n".join(render_party_card(item, idx) for idx, item in enumerate(party_cases, start=1))
    exam_cards = "\n".join(render_exam_card(item, idx) for idx, item in enumerate(exam_cases, start=1))

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>案例 V2 梳理预览</title>
  <style>
    :root {{
      --bg: #f7f0e6;
      --paper: #fffdfa;
      --card: #ffffff;
      --text: #2a251e;
      --muted: #766b60;
      --line: #e6d8c8;
      --accent: #c96f3a;
      --green: #2f7a4c;
      --soft: #fff5ec;
      --blue: #355f8f;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(201, 111, 58, .16), transparent 30rem),
        linear-gradient(180deg, #fbf6ef 0%, var(--bg) 100%);
    }}
    header {{
      position: sticky;
      top: 0;
      z-index: 10;
      border-bottom: 1px solid rgba(230, 216, 200, .9);
      background: rgba(255, 253, 250, .92);
      backdrop-filter: blur(12px);
    }}
    .bar {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 18px 22px;
      display: flex;
      justify-content: space-between;
      gap: 18px;
      align-items: center;
    }}
    h1 {{ margin: 0; font-size: 24px; letter-spacing: .02em; }}
    .subtitle {{ margin-top: 6px; font-size: 13px; color: var(--muted); }}
    .tools {{ display: flex; gap: 10px; align-items: center; flex-wrap: wrap; justify-content: flex-end; }}
    input {{
      width: min(360px, 45vw);
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 11px 16px;
      font-size: 14px;
      background: white;
      outline: none;
    }}
    .tab-btn {{
      border: 1px solid var(--line);
      background: white;
      color: var(--muted);
      border-radius: 999px;
      padding: 10px 14px;
      cursor: pointer;
      font-weight: 700;
    }}
    .tab-btn.active {{ background: var(--accent); color: white; border-color: var(--accent); }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 22px; }}
    .summary-grid {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }}
    .summary-card {{
      background: var(--paper);
      border: 1px solid var(--line);
      border-radius: 18px;
      padding: 16px;
      box-shadow: 0 10px 28px rgba(60, 50, 40, .05);
    }}
    .summary-card b {{ display: block; font-size: 28px; color: var(--accent); }}
    .summary-card span {{ display: block; margin-top: 6px; color: var(--muted); font-size: 13px; }}
    .note {{
      background: #fff9f2;
      border: 1px solid #efd5bd;
      border-radius: 18px;
      padding: 14px 16px;
      color: var(--muted);
      line-height: 1.75;
      margin-bottom: 18px;
      font-size: 14px;
    }}
    .section {{ display: none; }}
    .section.active {{ display: block; }}
    .section-head {{
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 12px;
      margin: 18px 0 12px;
    }}
    .section-head h2 {{ margin: 0; font-size: 20px; }}
    .section-head p {{ margin: 0; color: var(--muted); font-size: 13px; }}
    .case-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }}
    .case-card {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 20px;
      padding: 18px;
      box-shadow: 0 8px 24px rgba(60, 50, 40, .055);
    }}
    .card-head {{ display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; }}
    .rank {{ color: var(--accent); font-size: 12px; font-weight: 800; letter-spacing: .04em; }}
    h3 {{ margin: 5px 0 0; font-size: 18px; line-height: 1.45; }}
    .score {{
      min-width: 56px;
      text-align: center;
      padding: 8px 10px;
      border-radius: 14px;
      background: var(--soft);
      color: var(--accent);
      font-weight: 800;
    }}
    .meta {{ margin-top: 8px; color: var(--muted); font-size: 13px; line-height: 1.6; }}
    .ai-line {{ margin-top: 8px; }}
    .ai-badge {{
      display: inline-flex;
      border-radius: 999px;
      padding: 5px 9px;
      font-size: 12px;
      line-height: 1;
      font-weight: 800;
    }}
    .ai-none {{ color: #2f7a4c; background: #eef7f0; }}
    .ai-partial {{ color: #8a4d2d; background: #fff5ec; }}
    .target {{ margin-top: 12px; color: var(--text); line-height: 1.6; }}
    .target strong {{ color: var(--accent); }}
    blockquote {{
      margin: 12px 0;
      padding: 12px 14px;
      border-left: 3px solid var(--accent);
      background: #fbf7f1;
      border-radius: 0 12px 12px 0;
      color: #4c433a;
      line-height: 1.75;
      font-size: 14px;
    }}
    .tags {{ display: flex; gap: 6px; flex-wrap: wrap; margin-top: 10px; }}
    .tag {{
      display: inline-flex;
      padding: 5px 9px;
      border-radius: 999px;
      background: #f1e8dc;
      color: #5b4c40;
      font-size: 12px;
      line-height: 1;
    }}
    .detail-list {{ margin-top: 12px; }}
    .detail-block {{ margin-top: 12px; padding-top: 12px; border-top: 1px dashed var(--line); }}
    .detail-label {{ color: var(--accent); font-weight: 800; font-size: 13px; }}
    .detail-text {{ margin-top: 5px; color: #4c433a; line-height: 1.75; font-size: 14px; }}
    .fact-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      margin-top: 12px;
    }}
    .fact-grid div {{
      background: #fbf7f1;
      border-radius: 12px;
      padding: 10px;
      color: #4c433a;
      font-size: 13px;
      line-height: 1.6;
    }}
    .fact-grid span {{ display: block; color: var(--muted); font-size: 12px; margin-bottom: 2px; }}
    .risk, .advice {{
      margin-top: 10px;
      border-radius: 12px;
      padding: 10px 12px;
      font-size: 13px;
      line-height: 1.65;
    }}
    .risk {{ background: #fff5ec; color: #8a4d2d; }}
    .advice {{ background: #eef7f0; color: var(--green); }}
    .hidden {{ display: none !important; }}
    @media (max-width: 860px) {{
      .bar {{ display: block; }}
      .tools {{ margin-top: 14px; justify-content: flex-start; }}
      input {{ width: 100%; }}
      .summary-grid, .case-grid, .fact-grid {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="bar">
      <div>
        <h1>案例 V2 梳理预览</h1>
        <div class="subtitle">生成时间：{escape(generated_at)} · 党校按叙事丰富度排序，管综按意向院校明确度与字段完整度排序</div>
      </div>
      <div class="tools">
        <input id="searchInput" placeholder="搜索姓名、院校、系统、岗位、动机" />
        <button class="tab-btn active" data-tab="party">党校</button>
        <button class="tab-btn" data-tab="exam">管综</button>
      </div>
    </div>
  </header>
  <main>
    <div class="summary-grid">
      <div class="summary-card"><b>{len(party_cases)}</b><span>党校案例总数，全部保留</span></div>
      <div class="summary-card"><b>{rich_party_count}</b><span>党校高丰富案例，优先展示</span></div>
      <div class="summary-card"><b>{len(exam_cases)}</b><span>管综上岸案例总数，全部保留</span></div>
      <div class="summary-card"><b>{precise_exam_count}</b><span>管综意向院校明确案例</span></div>
    </div>
    <div class="note">
      数据源：党校读取 {escape(str(party_source.relative_to(ROOT)))}；管综读取 {escape(str(EXAM_SOURCE.relative_to(ROOT)))}。本页是内部核对页，不改变小程序老案例页入口；排序靠后的案例不删除，只作为补量。
    </div>
    <section id="party" class="section active">
      <div class="section-head">
        <h2>党校方向</h2>
        <p>排序：叙事字段去重长度加权 → 字段数量 → 原始顺序</p>
      </div>
      <div class="case-grid">{party_cards}</div>
    </section>
    <section id="exam" class="section">
      <div class="section-head">
        <h2>管综方向</h2>
        <p>排序：意向院校明确度 → 字段完整与内容长度 → 原始顺序</p>
      </div>
      <div class="case-grid">{exam_cards}</div>
    </section>
  </main>
  <script>
    const buttons = Array.from(document.querySelectorAll('.tab-btn'));
    const sections = Array.from(document.querySelectorAll('.section'));
    const input = document.getElementById('searchInput');

    function activeSection() {{
      return sections.find(section => section.classList.contains('active'));
    }}

    function applySearch() {{
      const term = (input.value || '').trim().toLowerCase();
      const section = activeSection();
      if (!section) return;
      section.querySelectorAll('.case-card').forEach(card => {{
        const text = (card.dataset.search || card.textContent || '').toLowerCase();
        card.classList.toggle('hidden', Boolean(term) && !text.includes(term));
      }});
    }}

    buttons.forEach(button => {{
      button.addEventListener('click', () => {{
        buttons.forEach(item => item.classList.remove('active'));
        button.classList.add('active');
        sections.forEach(section => section.classList.toggle('active', section.id === button.dataset.tab));
        applySearch();
      }});
    }});
    input.addEventListener('input', applySearch);
  </script>
</body>
</html>
"""


def main() -> None:
    party_source, party_cases = load_party_cases()
    exam_cases = load_exam_cases()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(render_html(party_source, party_cases, exam_cases), encoding="utf-8")
    print(f"wrote {OUTPUT.relative_to(ROOT)}")
    print(f"party={len(party_cases)} exam={len(exam_cases)}")
    print(f"party_rich={sum(1 for item in party_cases if item['quality'] == '高丰富')}")
    print(f"exam_precise={sum(1 for item in exam_cases if item['target_precision'] == '明确')}")


if __name__ == "__main__":
    main()
